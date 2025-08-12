
import pandas as pd
import numpy as np
from datetime import datetime
import os
import threading
import time
import streamlit as st
import matplotlib.pyplot as plt
import io
import plotly.express as px 
import plotly.graph_objects as go
from sklearn.linear_model import LinearRegression
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import webbrowser
import subprocess
import sys
import locale
import datetime as dt

meses = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}

hoy = pd.Timestamp.now()

mes_anio_actual = f"{meses[hoy.month]} {hoy.year}"

# --- Ajuste de límite de subida a 1 GB ---
config_path = os.path.expanduser("~/.streamlit/config.toml")
os.makedirs(os.path.dirname(config_path), exist_ok=True)
with open(config_path, "w") as f:
    f.write("[server]\nmaxUploadSize = 1024\n")

a = f"Análisis de Proveedores {mes_anio_actual}"

# Menú lateral para seleccionar página
pagina = st.sidebar.selectbox(
    "Selecciona la página",
    ["Configuración y carga", a]
)

# -------------------------
# Página 1: Configuración y carga
# -------------------------
if pagina == "Configuración y carga":
    st.title("⚙ Configuración y Carga de Archivos")
    st.markdown("Por favor, cargue todos los archivos necesarios. Las rutas sugeridas son para su referencia.")

    # Funciones de carga aquí (como tus cargar_dto, cargar_triage, etc.)


    # --- Funciones de carga ---
    @st.cache_data
    def cargar_dto(uploaded_file):
        dto = pd.read_excel(
            uploaded_file,
            sheet_name="DTO",
            usecols=["ID_SINIESTRO", "PROVEEDOR_POS", "PROVEEDOR_PROF", "ID_EMPRESA",
                    "RAZON_SOCIAL", "tipo_siniestro", "PROVEEDOR_ULTIMA_EDICION",
                    "PROVEEDOR_USUARIO_ULTIMA_EDICION", "FECHA_RADICACION", "NOMBRE_PROFESIONAL_POS",
                    "COMITE_INTER", "NOMBRE_COMITE_POS", "ORIGEN_POSITIVA", "EVENTO", "HEREDADO"]
        )
        return dto.rename(columns={
            "PROVEEDOR_POS": "PROVEEDOR_POS_TMO",
            "PROVEEDOR_ULTIMA_EDICION": "PROVEEDOR_ULTIMA_EDICION_TMO",
            "PROVEEDOR_USUARIO_ULTIMA_EDICION": "PROVEEDOR_USUARIO_ULTIMA_EDICION_TMO",
            "NOMBRE_PROFESIONAL_POS": "NOMBRE_PROFESIONAL_POS_TMO",
            "ORIGEN_POSITIVA": "ORIGEN_POSITIVA_TMO"
        })

    @st.cache_data
    def cargar_triage(uploaded_file):
        return pd.read_excel(uploaded_file, sheet_name="TRIAGE")

    @st.cache_data
    def cargar_convenio(uploaded_file):
        return pd.read_excel(uploaded_file, sheet_name="TOP CONVENIO")

    @st.cache_data
    def cargar_furat_el(uploaded_file):
        df = pd.read_excel(
            uploaded_file,
            header=1,
            usecols=["ID Siniestro", "Tipo de siniestro (AT o EL)", "Razón Social",
                    "Numero documento empresa", "Fecha de siniestro", "Hora del accidente",
                    "Fecha de Radicación", "Fecha de Muerte"]
        )
        df.rename(columns={
            "ID Siniestro": "ID_SINIESTRO",
            "Numero documento empresa": "NIT"
        }, inplace=True)
        return df

    @st.cache_data
    def cargar_ml(uploaded_file):
        return pd.read_csv(
            uploaded_file,
            encoding= "utf-8-sig",
            sep = ";",
            usecols=["CEDULA", "NOMBRE", "NIT", "EMPRESA", "NRO_SINIESTRO", "ESTADO",
                    "FECHA_RADICADO", "CARGUE", "FECHA_ACCIDENTE", "FECHA_ACTIVACION",
                    "ASIGNADO", "PRUEBA", "ORIGEN", "ALTO_COSTO", "FECHA_CARGUE",
                    "DICTAMEN", "ACTA", "EVENTO", "Dx", "PASO_COMITE", "ADICION_Dx",
                    "FECHA_DICTAMEN", "PRUEBA_SOBREVINIENTES", "FE_ERRATAS"]
        )

    # --- Uploader ---
    with st.expander("📂 Cargar archivos requeridos", expanded=True):
        archivo_dto = st.file_uploader(
            "DTO (TMO.xlsx - hoja 'DTO')\n📁 Ruta sugerida: BELISARIO S. A. S\\...\\Tmo Principal\\",
            type=["xlsx"]
        )
        archivo_triage = st.file_uploader(
            "TRIAGE (HIST_TRIAGE.xlsx - hoja 'TRIAGE')\n📁 Ruta sugerida: BELISARIO S. A. S\\...\\01 FUENTES DE INFORMACIÓN PRIMARIA\\",
            type=["xlsx"]
        )
        archivo_convenio = st.file_uploader(
            "CONVENIO (INFORMACION GENERAL.xlsx - hoja 'TOP CONVENIO')\n📁 Ruta sugerida: BELISARIO S. A. S\\...\\1.1. INFORMACIÓN GENERAL PROYECTO\\",
            type=["xlsx"]
        )
        archivo_furat = st.file_uploader(
            "FURAT (InformeGeneralAvisoSiniestroXXXX.xlsx)\n📁 Ruta sugerida: desde el gestor de reportes DOCUMN",
            type=["xlsx"]
        )
        archivo_ml = st.file_uploader(
            "ML (siniestros.xlsx)\n📁 Ruta sugerida: Desde el aplicativo ML",
            type=["csv"]
        )

    # --- Guardar en session_state ---
    if archivo_dto:
        st.session_state.DTO_renamed = cargar_dto(archivo_dto)
    if archivo_triage:
        st.session_state.TRIAGE_DATA = cargar_triage(archivo_triage)
    if archivo_convenio:
        st.session_state.CONVENIO = cargar_convenio(archivo_convenio)
    if archivo_furat:
        st.session_state.FURAT_EL = cargar_furat_el(archivo_furat)
    if archivo_ml:
        st.session_state.ML = cargar_ml(archivo_ml)



    if all(k in st.session_state for k in ["DTO_renamed", "TRIAGE_DATA", "CONVENIO", "FURAT_EL", "ML"]):
        st.success("✅ Todos los archivos han sido cargados correctamente. Procesando datos...")

        # Procesamiento de datos, usando las variables desde session_state
        DTO_renamed = st.session_state.DTO_renamed
        TRIAGE_DATA = st.session_state.TRIAGE_DATA
        CONVENIO = st.session_state.CONVENIO
        FURAT_EL = st.session_state.FURAT_EL
        ML = st.session_state.ML

        # --- Aquí va todo tu código de limpieza y uniones, por ejemplo: ---
        # Normalizar IDs
        FURAT_EL["ID_SINIESTRO"] = FURAT_EL["ID_SINIESTRO"].astype(str).str.strip()
        DTO_renamed["ID_SINIESTRO"] = DTO_renamed["ID_SINIESTRO"].astype(str).str.strip()
        TRIAGE_DATA["ID_SINIESTRO"] = TRIAGE_DATA["ID_SINIESTRO"].astype(str).str.strip()

        # Copias para trabajar
        furat1 = FURAT_EL.copy()
        furat = FURAT_EL.copy()
        dto = DTO_renamed.copy()
        convenio = CONVENIO.copy()
        ml = ML.copy()
        ml = ml.rename(columns={"NRO_SINIESTRO": "ID_SINIESTRO"})
        ml["ID_SINIESTRO"] = ml["ID_SINIESTRO"].astype(str).str.strip()
        triage = TRIAGE_DATA.copy()

        
        furat.rename(columns={"ID Siniestro": "ID_SINIESTRO", "Numero documento empresa": "NIT"}, inplace=True)
        furat1.rename(columns={"ID Siniestro": "ID_SINIESTRO", "Numero documento empresa": "NIT"}, inplace=True)
        
        furat["Fecha de Radicación"] = pd.to_datetime(
        furat["Fecha de Radicación"],
        errors='coerce',
        dayfirst=True,
        format='mixed')
        
        furat1["Fecha de Radicación"] = pd.to_datetime(
        furat1["Fecha de Radicación"],
        errors='coerce',
        dayfirst=True,
        format='mixed')

        dto["FECHA_RADICACION"] = pd.to_datetime(
            dto["FECHA_RADICACION"],
            errors='coerce',
            dayfirst=True,
            format='mixed'
        )

        # ========================================
        # UNIONES DE TABLAS BASE PARA ANÁLISIS
        # ========================================

        # 1. Unir DTO por 'ID_SINIESTRO'
        furat = furat.merge(dto, how="left", on="ID_SINIESTRO")

        # 2. Unir CONVENIO por 'NIT'
        furat = furat.merge(
            convenio[["NIT", "NUEVA ASIGNACIÓN PROVEEDOR DE CALIFICACIÓN ", "FECHA DE INACTIVACION / RETIRO"]],
            how="left",
            on="NIT"
        )

        # 3. Unir TRIAGE por 'ID_SINIESTRO'
        furat = furat.merge(
            triage[["ID_SINIESTRO", "CALIF"]],
            how="left",
            on="ID_SINIESTRO"
        )

        # 4. Unir ML por 'ID_SINIESTRO'
        furat = furat.merge(
            ml[["ID_SINIESTRO", "ESTADO"]],
            how="left",
            on="ID_SINIESTRO"
        )

        # Limpiar caracteres especiales en fechas
        furat["FECHA DE INACTIVACION / RETIRO"] = furat["FECHA DE INACTIVACION / RETIRO"].replace("\xa0", np.nan)


        # ========================================
        # CAMBIOS Y FILTROS
        # ========================================

        furat = furat[(furat["ID_SINIESTRO"] != "0")]


        furat = furat[(furat["Tipo de siniestro (AT o EL)"] != "EL")]

        furat = furat[
            ~furat["NOMBRE_COMITE_POS"].str.contains("C. de cargue y trazabilidad|HEREDADAS TRAZA", case=False, na=False)
        ]

        furat = furat[furat["tipo_siniestro"] != "EP"]
        # ========================================
        # PROCESAMIENTO DE FURAT Y DTO UNIFICADO
        # ========================================

        # --------
        # FURAT
        # --------
        furat["ID_SINIESTRO"] = furat["ID_SINIESTRO"].astype(str).str.strip()
        furat["PROVEEDOR1"] = furat["ID_SINIESTRO"].str[-1]
        furat["PROVEEDOR2"] = None
        furat["PROVEEDORTRIAGE"] = None
        furat["PROVEEDOR"] = None

        # Asignación PROVEEDOR2
        furat.loc[(furat["PROVEEDOR1"].isin(["1", "2", "3", "4", "5"])) &
                (furat["NUEVA ASIGNACIÓN PROVEEDOR DE CALIFICACIÓN "] == "Belisario SAS"), "PROVEEDOR2"] = "BELISARIO"
        furat.loc[(furat["PROVEEDOR1"].isin(["1", "2", "3", "4", "5"])) &
                (furat["NUEVA ASIGNACIÓN PROVEEDOR DE CALIFICACIÓN "] != "Belisario SAS"), "PROVEEDOR2"] = "GESTAR"
        furat.loc[(furat["PROVEEDOR1"].isin(["6", "7", "8", "9", "0"])) &
                (furat["NUEVA ASIGNACIÓN PROVEEDOR DE CALIFICACIÓN "] == "Belisario SAS"), "PROVEEDOR2"] = "BELISARIO"
        furat.loc[(furat["PROVEEDOR1"].isin(["6", "7", "8", "9", "0"])) &
                (furat["NUEVA ASIGNACIÓN PROVEEDOR DE CALIFICACIÓN "]
                .isin(["Unión temporal MDL 2023", "EMPRESA CONFLICTO DE INTERES BELISARIO "])), "PROVEEDOR2"] = "GESTAR"

        # PROVEEDORTRIAGE
        furat.loc[(furat["CALIF"] == "TRIAGE") & (furat["PROVEEDOR2"] == "BELISARIO"), "PROVEEDORTRIAGE"] = "BELISARIO_TRIAGE"
        furat.loc[(furat["CALIF"] == "TRIAGE") & (furat["PROVEEDOR2"] == "GESTAR"), "PROVEEDORTRIAGE"] = "GESTAR_TRIAGE"

        # PROVEEDOR Final
        furat.loc[(furat["CALIF"].isna()) & (furat["PROVEEDOR2"] == "BELISARIO"), "PROVEEDOR"] = "BVS"
        furat.loc[(furat["CALIF"].isna()) & (furat["PROVEEDOR2"] == "GESTAR"), "PROVEEDOR"] = "Gestar"
        furat.loc[(furat["CALIF"] == "TRIAGE") & (furat["PROVEEDOR1"].isin(["1", "2", "3", "4", "5"])), "PROVEEDOR"] = "Gestar_TRIAGE"
        furat.loc[(furat["CALIF"] == "TRIAGE") & (furat["PROVEEDOR1"].isin(["6", "7", "8", "9", "0"])), "PROVEEDOR"] = "BVS_TRIAGE"
        furat.loc[furat["ESTADO"].notna(), "PROVEEDOR"] = "BVS"
        furat["FUENTE"] = "FURAT"
        

        # --------
        # DTO NUEVOS
        # --------
        
        dto.rename(columns={"ID_EMPRESA": "NIT", "RAZON_SOCIAL": "Razón Social"}, inplace=True)
        dto = dto.merge(convenio[["NIT", "NUEVA ASIGNACIÓN PROVEEDOR DE CALIFICACIÓN ", "FECHA DE INACTIVACION / RETIRO"]], how="left", on="NIT")
        dto = dto.merge(triage[["ID_SINIESTRO", "CALIF"]], how="left", on="ID_SINIESTRO")
        dto = dto.merge(ml[["ID_SINIESTRO", "ESTADO"]], how="left", on="ID_SINIESTRO")
        dto = dto.merge(furat1[["ID_SINIESTRO", "Fecha de Muerte"]], how="left", on="ID_SINIESTRO")

        dto["FECHA DE INACTIVACION / RETIRO"] = dto["FECHA DE INACTIVACION / RETIRO"].replace("\xa0", np.nan)

        # Filtro temporal
        inicio = furat["Fecha de Radicación"].min()
        fin = furat["Fecha de Radicación"].max()

                
        filtro = (~dto["NOMBRE_COMITE_POS"].str.contains("C. de cargue y trazabilidad|HEREDADAS TRAZA", case=False, na=False) &
                (dto["FECHA_RADICACION"] >= inicio) & (dto["FECHA_RADICACION"] <= fin))
        df_filtrado = dto[filtro].copy()
        df_filtrado["ID_SINIESTRO"] = df_filtrado["ID_SINIESTRO"].astype(str).str.strip()

        # Extraer nuevos
        ids_furat = set(furat["ID_SINIESTRO"].unique())
        df_nuevos = df_filtrado[~df_filtrado["ID_SINIESTRO"].isin(ids_furat)].copy()
        df_nuevos["Tipo de siniestro (AT o EL)"] = df_nuevos["ID_SINIESTRO"].map(
            furat.drop_duplicates(subset="ID_SINIESTRO").set_index("ID_SINIESTRO")["Tipo de siniestro (AT o EL)"]
        )
        df_nuevos["Fecha de Muerte"] = None
        df_nuevos["Hora del accidente"] = None
        df_nuevos["PROVEEDOR1"] = df_nuevos["ID_SINIESTRO"].str[-1]
        df_nuevos["PROVEEDOR2"] = None

        # Asignaciones mismas que FURAT
        mismo_proceso = lambda df: None  

        # Usamos mismo código de asignación que FURAT para df_nuevos
        # (puede compactarse en función si quieres luego)
        df_nuevos.loc[(df_nuevos["PROVEEDOR1"].isin(["1", "2", "3", "4", "5"])) &
                    (df_nuevos["NUEVA ASIGNACIÓN PROVEEDOR DE CALIFICACIÓN "] != "Belisario SAS"), "PROVEEDOR2"] = "GESTAR"
        df_nuevos.loc[(df_nuevos["PROVEEDOR1"].isin(["6", "7", "8", "9", "0"])) &
                    (df_nuevos["NUEVA ASIGNACIÓN PROVEEDOR DE CALIFICACIÓN "] == "Belisario SAS"), "PROVEEDOR2"] = "BELISARIO"
        df_nuevos.loc[(df_nuevos["PROVEEDOR1"].isin(["6", "7", "8", "9", "0"])) &
                    (df_nuevos["NUEVA ASIGNACIÓN PROVEEDOR DE CALIFICACIÓN "]
                    .isin(["Unión temporal MDL 2023", "EMPRESA CONFLICTO DE INTERES BELISARIO "])), "PROVEEDOR2"] = "GESTAR"
        df_nuevos.loc[(df_nuevos["PROVEEDOR1"].isin(["1", "2", "3", "4", "5"])) &
                    (df_nuevos["NUEVA ASIGNACIÓN PROVEEDOR DE CALIFICACIÓN "] == "Belisario SAS"), "PROVEEDOR2"] = "BELISARIO"

        df_nuevos["PROVEEDORTRIAGE"] = None
        df_nuevos.loc[(df_nuevos["PROVEEDOR2"] == "BELISARIO") & (df_nuevos["CALIF"] == "TRIAGE"), "PROVEEDORTRIAGE"] = "BELISARIO_TRIAGE"
        df_nuevos.loc[(df_nuevos["PROVEEDOR2"] == "GESTAR") & (df_nuevos["CALIF"] == "TRIAGE"), "PROVEEDORTRIAGE"] = "GESTAR_TRIAGE"

        # Final

        df_nuevos["PROVEEDOR"] = None
        df_nuevos.loc[(df_nuevos["PROVEEDOR2"] == "BELISARIO") & (df_nuevos["CALIF"].isna()), "PROVEEDOR"] = "BVS"
        df_nuevos.loc[(df_nuevos["PROVEEDOR2"] == "GESTAR") & (df_nuevos["CALIF"].isna()), "PROVEEDOR"] = "Gestar"
        df_nuevos.loc[(df_nuevos["PROVEEDORTRIAGE"] == "BELISARIO_TRIAGE"), "PROVEEDOR"] = "BVS"
        df_nuevos.loc[(df_nuevos["PROVEEDORTRIAGE"] == "GESTAR_TRIAGE"), "PROVEEDOR"] = "Gestar"
        df_nuevos.loc[(df_nuevos["CALIF"] == "TRIAGE") & (df_nuevos["PROVEEDOR1"].isin(["1", "2", "3", "4", "5"])), "PROVEEDOR"] = "Gestar_TRIAGE"
        df_nuevos.loc[(df_nuevos["CALIF"] == "TRIAGE") & (df_nuevos["PROVEEDOR1"].isin(["6", "7", "8", "9", "0"])), "PROVEEDOR"] = "BVS_TRIAGE"
        df_nuevos.loc[df_nuevos["ESTADO"].notna(), "PROVEEDOR"] = "BVS"
        df_nuevos["FUENTE"] = "TMO"

        # --------
        # UNION FINAL
        # --------
        
        
        columnas_comunes = furat.columns.intersection(df_nuevos.columns)
        furat = pd.concat([furat, df_nuevos])

        furat["PROVEEDOR"] = furat["PROVEEDOR"].replace({None: np.nan})

        furat.loc[
            (furat["PROVEEDOR"].isna()) & (furat["PROVEEDOR1"].isin(["6", "7", "8", "9", "0"])),
            "PROVEEDOR"
        ] = "BVS"

        furat.loc[
            (furat["PROVEEDOR"].isna()) & (furat["PROVEEDOR1"].isin(["1", "2", "3", "4", "5"])),
            "PROVEEDOR"
        ] = "Gestar"


        furat.loc[(furat["PROVEEDORTRIAGE"] == "BELISARIO_TRIAGE") , "PROVEEDOR"] = "BVS_TRIAGE"
        furat.loc[(furat["PROVEEDORTRIAGE"] == "GESTAR_TRIAGE") , "PROVEEDOR"] = "Gestar_TRIAGE"



        furat.loc[(furat["Razón Social"].str.contains("SECRETARIA")) , "PROVEEDOR"] = "Gestar"

        furat = furat[(furat["Fecha de Muerte"] == "No Aplica") | (furat["Fecha de Muerte"].isna())]
        


        #899999061 secretaria - enumerar
        id_secretaria = "899999061"
        mask_secretaria = furat["NIT"] == id_secretaria
        furat.loc[mask_secretaria, "NIT_ENUM"] = [f"{id_secretaria}_{i+1}" for i in range(mask_secretaria.sum())]
        furat.loc[~mask_secretaria, "NIT_ENUM"] = furat.loc[~mask_secretaria, "NIT"]
        
        

        furat["fecha_radicacion"] = furat["Fecha de Radicación"].replace("", np.nan).fillna(furat["FECHA_RADICACION"])
        
        if "fecha_radicacion" in furat.columns:
            col_temp = furat.pop("fecha_radicacion")
            furat.insert(24, "fecha_radicacion", col_temp)
            
        furat["fecha_radicacion"] = pd.to_datetime(
            furat["fecha_radicacion"],
            errors='coerce',
            dayfirst=True,
            format='mixed'
        )


        df_filtrado = furat.copy()
        st.session_state["df_filtrado"] = df_filtrado
        st.session_state["furat"] = furat


        # Al final guarda el DataFrame resultante en session_state para usar en la segunda página:
        st.session_state["df_procesado"] = furat

        st.info("Datos procesados y listos para análisis. Ahora puede ir a la página de análisis.")

    else:
        st.warning("⚠ Debe cargar todos los archivos antes de continuar.")



#####################################3

# -------------------------
# Página 2: Análisis de Proveedores
# -------------------------
elif pagina == a:
    st.title(a)
    if "df_filtrado" not in st.session_state:
        st.warning("⚠ Por favor, primero cargue y procese los archivos en la página 'Configuración y carga'.")
    else:
        # -------------------------
        # Configuración de la página
        # -------------------------
        
        df_filtrado = st.session_state["df_filtrado"]
        furat = st.session_state["furat"]


        st.set_page_config(page_title= a, layout="wide")

        # Título con estilo personalizado
        fecha_min = df_filtrado["fecha_radicacion"].min().strftime("%d %B %Y")
        fecha_max = df_filtrado["fecha_radicacion"].max().strftime("%d %B %Y")
        st.markdown(f"""
            <style>
                .main-title {{
                    font-size: 40px;
                    font-weight: 700;
                    text-align: center;
                    padding: 10px;
                    background: linear-gradient(to right, #5d5f85, #28405c);
                    color: white;
                    border-radius: 12px;
                    box-shadow: 0 4px 10px rgba(0,0,0,0.2);
                    margin-bottom: 10px;
                }}
                .date-range {{
                    font-size: 18px;
                    font-weight: 500;
                    text-align: center;
                    color: #1a3d5d;
                    background-color: #d6e9f9;
                    border-radius: 8px;
                    padding: 6px 12px;
                    display: inline-block;
                    box-shadow: 0 2px 6px rgba(0,0,0,0.1);
                    margin-bottom: 30px;
                }}
                .date-container {{
                    text-align: center;
                }}
            </style>
            <div class="main-title">{a}</div>
            <div class="date-container">
                <div class="date-range">DATOS desde <b>{fecha_min}</b> HASTA <b>{fecha_max}</b></div>
            </div>
        """, unsafe_allow_html=True)
        # -------------------------
        # 🧩 Filtros Interactivos mejorados
        # -------------------------
        st.sidebar.markdown("""
            <style>
                .filtro-box {
                    background-color: #f1f3f6;
                    padding: 20px 15px;
                    border-radius: 10px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                    margin-bottom: 20px;
                }
                .filtro-box h4 {
                    text-align: center;
                    margin-bottom: 10px;
                    color: #333333;
                }
                .stCheckbox > label {
                    font-weight: 500;
                    font-size: 15px;
                    color: #333333;
                }
            </style>
            <div class='filtro-box'>
                <h4>🔎 Filtros</h4>
            </div>
        """, unsafe_allow_html=True)

        with st.sidebar.container():
            filtro_solo_triage = st.checkbox("✅ Solo TRIAGE")
            filtro_sin_estado = st.checkbox("⚠️ Sin estado belisario")
            filtro_sin_proveedor = st.checkbox("🚫 Sin convenio")

        # Copiar el DataFrame original
        df_filtrado = furat.copy()

        # Aplicar filtros
        if filtro_solo_triage:
            df_filtrado = df_filtrado[df_filtrado["PROVEEDORTRIAGE"].notna()]
        if filtro_sin_estado:
            df_filtrado = df_filtrado[df_filtrado["ESTADO"].isna()]
        if filtro_sin_proveedor:
            df_filtrado = df_filtrado[df_filtrado["NUEVA ASIGNACIÓN PROVEEDOR DE CALIFICACIÓN "].isna()]



        # -------------------------
        # 🔢 Métricas principales
        # -------------------------
        col1, col2, col3 = st.columns(3)

        # Casos Totales
        with col1:
            st.markdown(f"""
            <div style="background-color:#2E86C1; padding:20px; border-radius:15px; text-align:center;
            color:white; box-shadow: 2px 2px 5px rgba(0,0,0,0.2)">
                <h4>Casos Totales</h4>
                <h2>{len(df_filtrado["PROVEEDOR"])}</h2>
            </div>
            """, unsafe_allow_html=True)

        # TRIAGE
        with col2:
            total_triage = df_filtrado["CALIF"].eq("TRIAGE").sum()
            st.markdown(f"""
            <div style="background-color:#28B463; padding:20px; border-radius:15px; text-align:center;
            color:white; box-shadow: 2px 2px 5px rgba(0,0,0,0.2)">
                <h4>TRIAGE</h4>
                <h2>{total_triage}</h2>
            </div>
            """, unsafe_allow_html=True)


        # Convenios Belisario
        with col3:
            total_convenios_bvs = df_filtrado[(df_filtrado["PROVEEDOR"] == "BVS") &
                                            (df_filtrado["NUEVA ASIGNACIÓN PROVEEDOR DE CALIFICACIÓN "] == "Belisario SAS")].shape[0]
            st.markdown(f"""
            <div style="background-color:#CA6F1E; padding:20px; border-radius:15px; text-align:center;
            color:white; box-shadow: 2px 2px 5px rgba(0,0,0,0.2)">
                <h4>Convenios</h4>
                <h2>{total_convenios_bvs}</h2>
            </div>
            """, unsafe_allow_html=True)
            
            
        # -------------------------
        # 📅 Filtro de fechas
        # -------------------------
       
        # 🎨 Estilo CSS más moderno
        st.markdown("""
        <style>
        .date-filter-box {
            background-color: white;
            padding: 15px 20px;
            border-radius: 12px;
            margin-top: 15px;
            border: 1px solid #E5E7EB;
            box-shadow: 0 2px 5px rgba(0,0,0,0.05);
        }
        .date-filter-title {
            font-size: 17px;
            font-weight: bold;
            color: #2563EB;
            margin-bottom: 8px;
            display: flex;
            align-items: center;
        }
        .date-filter-title::before {
            content: "📆";
            margin-right: 6px;
        }
        </style>
        """, unsafe_allow_html=True)

        with st.container():
            st.markdown('<div class="date-filter-box">', unsafe_allow_html=True)
            st.markdown('<div class="date-filter-title">Filtrar por Fecha de Radicación</div>', unsafe_allow_html=True)

            # Obtener fecha mínima y máxima
            min_fecha = df_filtrado["fecha_radicacion"].min().date()
            max_fecha = df_filtrado["fecha_radicacion"].max().date()

            # Selector de fechas (permite uno o dos valores)
            rango_fechas = st.date_input(
                "",
                value=(min_fecha, max_fecha),
                min_value=min_fecha,
                max_value=max_fecha
            )

            st.markdown('</div>', unsafe_allow_html=True)

        # 🛠 Manejo de selección única o doble
        if isinstance(rango_fechas, tuple) and len(rango_fechas) == 2:
            fecha_inicio, fecha_fin = rango_fechas
        elif isinstance(rango_fechas, dt.date):
            fecha_inicio = fecha_fin = rango_fechas
        else:
            st.warning("Por favor selecciona al menos una fecha.")
            st.stop()

        # 📊 Filtrar el DataFrame
        df_filtrado = df_filtrado[
            (df_filtrado["fecha_radicacion"] >= pd.to_datetime(fecha_inicio)) &
            (df_filtrado["fecha_radicacion"] <= pd.to_datetime(fecha_fin))
        ]
        # -------------------------
        # 📊 Gráficas tipo dash
        # -------------------------

        # --- Distribución final por PROVEEDOR (pie + tabla) ---    
        with st.container():
            st.subheader("")

            col1, col2 = st.columns([1, 2]) 

        # ⬅️ GRÁFICA A LA IZQUIERDA
            with col1:
                df_prov = df_filtrado[
                df_filtrado["PROVEEDOR"].isin(["Gestar", "BVS"]) &
                (df_filtrado["NUEVA ASIGNACIÓN PROVEEDOR DE CALIFICACIÓN "].isna())
            ]
            
                proveedor_counts = df_prov["PROVEEDOR"].value_counts().reset_index()
                proveedor_counts.columns = ["PROVEEDOR", "N"]

                colores_proveedor = {
                    "BVS": "#0b8cce",         
                    "Gestar": "#f29559",     
                    "Otro": "#aab2bd"        
                }

                colores = [colores_proveedor.get(p, "#aab2bd") for p in proveedor_counts["PROVEEDOR"]]

                fig = px.pie(
                    proveedor_counts,
                    names="PROVEEDOR",
                    values="N",
                    hole=0
                )

                fig.update_traces(
                    textposition="inside",
                    textinfo="label+percent",
                    textfont=dict(color="white", size=14),
                    marker=dict(colors=colores, line=dict(color='#000000', width=1))
                )

                fig.update_layout(
                    height=300,
                    margin=dict(t=10, b=10, l=10, r=10),
                    showlegend=True,
                    legend=dict(
                        orientation="h",
                        yanchor="bottom",
                        y=-0.2,
                        xanchor="center",
                        x=0.5,
                        font=dict(size=12)
                    ),
                    paper_bgcolor="rgba(0,0,0,0)",  
                    plot_bgcolor="rgba(0,0,0,0)",
                )

                st.markdown("""
                <div style="
                    background: linear-gradient(135deg, #e0eafc, #cfdef3); 
                    padding: 25px; 
                    border-radius: 15px; 
                    box-shadow: 0 4px 15px rgba(0, 0, 0, 0.15); 
                    margin-top: 0px;
                    max-width: 500px;
                    margin-right: auto;
                    text-align: center;
                ">
                    <h4 style="margin: 0; padding-bottom: 10px; color: #1f3a93;">Distribución por proveedor</h4>
                """, unsafe_allow_html=True)

                st.plotly_chart(fig, use_container_width=False)

                st.markdown("</div>", unsafe_allow_html=True)

            # TABLA DERECHA
            with col2:
                st.markdown("""
                <div style="
                    background: linear-gradient(135deg, #0b8cce, #0d2f66);
                    padding: 20px;
                    border-radius: 15px; 
                    max-width: 1200px;
                    box-shadow: 0 4px 50px rgba(0, 0, 0, 0.1);
                    margin-top: 5px;
                    text-align: center;
                ">
                    <h4 style="color: white; margin: 0;">Casos semanales por proveedor</h4>
                </div>
                """, unsafe_allow_html=True)

                st.markdown("<div style='height: 20px;'></div>", unsafe_allow_html=True)
                

                df_temp = df_filtrado.copy()
                df_temp["Semana"] = pd.cut(
                    df_temp["fecha_radicacion"].dt.day,
                    bins=[0, 7, 14, 21, 28, 31],
                    labels=["Semana 1 (1-7)", "Semana 2 (8-14)", "Semana 3 (15-21)", "Semana 4 (22-28)", "Semana 5 (29-31)"],
                    include_lowest=True
                )
                
                dia_min = df_temp["fecha_radicacion"].dt.day.min()
                dia_max = df_temp["fecha_radicacion"].dt.day.max()

                # Calcular el rango de días
                num_dias = dia_max - dia_min + 1

                # Definir tamaño de intervalo dinámico
                if num_dias <= 7:
                    intervalo = 1   # un bin por día
                elif num_dias <= 14:
                    intervalo = 2   # grupos de 2 días
                elif num_dias <= 21:
                    intervalo = 3   # grupos de 3 días
                else:
                    intervalo = 7   # grupos semanales

                # Crear los bins dinámicos
                bins = list(range(dia_min, dia_max + intervalo, intervalo))
                labels = [f"Días {bins[i]}-{bins[i+1]-1}" for i in range(len(bins)-1)]

                # Asignar la columna "Semana" pero dinámica
                df_temp["Semana"] = pd.cut(
                    df_temp["fecha_radicacion"].dt.day,
                    bins=bins,
                    labels=labels,
                    include_lowest=True
                )

                

                tabla_resumen = pd.DataFrame()
                tabla_resumen["BVS"] = df_temp[((df_temp["PROVEEDOR"] == "BVS") & (df_temp["PROVEEDORTRIAGE"].isna()))].groupby("Semana").size()
                tabla_resumen["TRIAGE BVS"] = df_temp[((df_temp["PROVEEDOR"] == "BVS_TRIAGE") )].groupby("Semana").size()
                tabla_resumen["Gestar"] = df_temp[((df_temp["PROVEEDOR"] == "Gestar") & (df_temp["PROVEEDORTRIAGE"].isna()))].groupby("Semana").size()
                tabla_resumen["TRIAGE GTR"] = df_temp[((df_temp["PROVEEDOR"] == "Gestar_TRIAGE") )].groupby("Semana").size()

                tabla_resumen = tabla_resumen.fillna(0).astype(int)
                tabla_resumen["Total general"] = tabla_resumen.sum(axis=1)
                total_global = tabla_resumen["Total general"].sum()

                tabla_resumen["Total porcentual"] = round(tabla_resumen["Total general"] / total_global * 100, 2)

                total_final = pd.DataFrame(tabla_resumen.sum(numeric_only=True)).T
                total_final.index = ["Total general"]

                tabla_final = pd.concat([tabla_resumen, total_final])


                format_dict = {col: "{: }" for col in tabla_final.columns if col != "Total porcentual"}
                format_dict["Total porcentual"] = "{:.2f}%"
                
                styled_table = tabla_final.style \
                    .format(format_dict) \
                    .set_properties(**{
                        'text-align': 'center',
                        'font-size': '13px',
                    }) \
                    .set_table_styles([
                        {"selector": "thead th", "props": [("background-color", "#0b8cce"), ("color", "white"), ("font-size", "14px")]},
                        {"selector": "tbody td", "props": [("border", "1px solid #ccc")]},
                    ]) \
                    .apply(lambda x: ['background-color: #e6f2fa' if i % 2 == 0 else 'background-color: white' for i in range(len(x))], axis=0)


                st.write(styled_table)

        # -------------------------
        # 📈 Barra diaria de radicación BVS  (Julio)
        # -------------------------


        df_bvs = df_filtrado[df_filtrado["PROVEEDOR"] == "BVS"].copy()
        df_bvs["Día"] = df_bvs["fecha_radicacion"].dt.day
        df_bvs["DíaSemana"] = df_bvs["fecha_radicacion"].dt.dayofweek  # 0 = Lunes, 6 = Domingo

        # Lunes a Viernes (0 a 4)
        df_laborales = df_bvs[df_bvs["DíaSemana"] <= 4]

        # Sábados y Domingos (5 y 6)
        df_finde = df_bvs[df_bvs["DíaSemana"] >= 5]

        conteo_laborales = df_laborales["Día"].value_counts().sort_index()
        conteo_finde = df_finde["Día"].value_counts().sort_index()


        def generar_grafico_radicaciones(dias_conteo, titulo):
            # Filtramos solo días con datos
            dias_validos = dias_conteo[dias_conteo > 0].index.tolist()
            radicaciones = dias_conteo[dias_conteo > 0].values

            # Preparamos datos para regresión
            dias_np = np.array(dias_validos).reshape(-1, 1)
            modelo = LinearRegression()
            modelo.fit(dias_np, radicaciones)
            tendencia = modelo.predict(dias_np)

            fig = go.Figure()

            # Barras
            fig.add_trace(go.Bar(
                x=dias_validos,
                y=radicaciones,
                name='Casos diarios',
                marker=dict(color='#0b8cce'),
                hoverinfo='x+y',
            ))

            # Línea de tendencia
            fig.add_trace(go.Scatter(
                x=dias_validos,
                y=tendencia,
                mode='lines',
                name='Tendencia',
                line=dict(color='blue', width=2, dash='dash')
            ))

            fig.update_layout(
                title=titulo,
                xaxis=dict(title='Día del mes', dtick=1, tickmode='linear', tickfont=dict(size=16)),
                yaxis=dict(title='Número de casos', tickfont=dict(size=16)),
                height=450,
                template='plotly_white',
                legend=dict(orientation="h", y=1.02, x=0.5, xanchor="center", yanchor="bottom"),
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)"
            )


            return fig



        fig_laborales = generar_grafico_radicaciones(conteo_laborales, "📅 Lunes a Viernes")
        fig_finde = generar_grafico_radicaciones(conteo_finde, "🗓️ Sábado y Domingo")

        # -------------------------
        # Estilo encabezado
        # -------------------------

        st.markdown("""
            <div style="background: linear-gradient(to right, #0b8cce, #0d2f66); 
                        padding: 25px; 
                        border-radius: 15px; 
                        box-shadow: 0 4px 10px rgba(0, 0, 0, 0.15); 
                        margin-top: 30px;
                        margin-bottom: 20px;">
                <h4 style="text-align: center; color: #ffffff; margin-bottom: 0px; font-size: 30px;">
                    Radicaciones Diarias - BVS (Julio)
                </h4>
            </div>
        """, unsafe_allow_html=True)

        # -------------------------
        # Gráfica 1: Lunes a Viernes
        # -------------------------
        st.markdown("""
            <div style="background-color: #f0f8ff; padding: 20px; border-radius: 15px; margin-bottom: 20px;">
                <h5 style="color: #0b4c6f;">🔹 Radicaciones entre semana (Lunes a Viernes)</h5>
            </div>
        """, unsafe_allow_html=True)
        st.plotly_chart(fig_laborales, use_container_width=True)

        # -------------------------
        # Gráfica 2: Fin de Semana
        # -------------------------
        st.markdown("""
            <div style="background-color: #fff5f0; padding: 20px; border-radius: 15px; margin-bottom: 20px;">
                <h5 style="color: #a63603;">🔹 Radicaciones fin de semana (Sábado y Domingo)</h5>
            </div>
        """, unsafe_allow_html=True)
        st.plotly_chart(fig_finde, use_container_width=True)


        # -------------------------
        # 💾 Exportación a Excel
        # -------------------------

        @st.cache_data
        def convertir_a_excel(df):
            output = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Datos"

            color_map = {
                "amarillo": ["ID_SINIESTRO", "Tipo de siniestro (AT o EL)", "Fecha de siniestro", "Fecha de Radicación", "NIT", "Razón Social","Fecha de Muerte",
                            "Hora del accidente"],
                "azul": ["PROVEEDOR_ULTIMA_EDICION_TMO", "PROVEEDOR_USUARIO_ULTIMA_EDICION_TMO", "PROVEEDOR_PROF","tipo_siniestro",
                        "ID_EMPRESA","RAZON_SOCIAL","ORIGEN_POSITIVA_TMO", "NOMBRE_PROFESIONAL_POS_TMO", "PROVEEDOR_POS_TMO",
                        "HEREDADO","EVENTO","COMITE_INTER","NOMBRE_COMITE_POS","FECHA_RADICACION"],
                "naranja": ["NUEVA ASIGNACIÓN PROVEEDOR DE CALIFICACIÓN ", "FECHA DE INACTIVACION / RETIRO"],
                "verde": ["CALIF", "ESTADO","FUENTE"],
                "salmon": ["PROVEEDOR1", "PROVEEDOR2", "PROVEEDORTRIAGE", "PROVEEDOR"],
                "azul_oscuro": ["fecha_radicacion"]
            }

            fill_colors = {
                "amarillo": PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid"),
                "azul": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
                "naranja": PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid"),
                "verde": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
                "salmon": PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"),
                "azul_oscuro": PatternFill(start_color="0b8cce", end_color="0b8cce", fill_type="solid")
            }

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
                    if r_idx == 1:
                        col_name = value
                        for color, columnas in color_map.items():
                            if col_name in columnas:
                                cell.fill = fill_colors[color]

            for column_cells in ws.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2

            wb.save(output)
            output.seek(0)
            return output.getvalue()

        def aplicar_estilo(df):
            color_map = {
                "amarillo": ["ID_SINIESTRO", "Tipo de siniestro (AT o EL)", "Fecha de siniestro", "Fecha de Radicación", "NIT", "Razón Social",
                            "Hora del accidente","Fecha de Muerte"],
                "azul": ["PROVEEDOR_ULTIMA_EDICION_TMO", "PROVEEDOR_USUARIO_ULTIMA_EDICION_TMO", "PROVEEDOR_PROF", "ORIGEN_POSITIVA_TMO","ID_EMPRESA","RAZON_SOCIAL",
                        "tipo_siniestro",
                        "NOMBRE_PROFESIONAL_POS_TMO", "PROVEEDOR_POS_TMO", "HEREDADO", "EVENTO", "COMITE_INTER", "NOMBRE_COMITE_POS","FECHA_RADICACION"],
                "naranja": ["NUEVA ASIGNACIÓN PROVEEDOR DE CALIFICACIÓN ", "FECHA DE INACTIVACION / RETIRO"],
                "verde": ["CALIF", "ESTADO", "FUENTE"],
                "salmon": ["PROVEEDOR1", "PROVEEDOR2", "PROVEEDORTRIAGE", "PROVEEDOR"],
                "azul_oscuro": ["fecha_radicacion"]
            }

            color_hex = {
                "amarillo": "#FFFACD",
                "azul": "#ADD8E6",
                "naranja": "#FFD580",
                "verde": "#90EE90",
                "salmon": "#FFA07A",
                "azul_oscuro": "#0b8cce"
            }

            styles = []
            for color, cols in color_map.items():
                for col in cols:
                    if col in df.columns:
                        col_idx = df.columns.get_loc(col)
                        styles.append({
                            "selector": f"th.col_heading.level0.col{col_idx}",
                            "props": [("background-color", color_hex[color]), ("font-weight", "bold")]
                        })

            # Solo devuelve estilo, sin limitar aquí
            styled_df = df.style.set_table_styles(styles)
            return styled_df

        # Mostrar en Streamlit
        if 'df_filtrado' in locals() or 'df_filtrado' in globals():
            st.markdown("### Previsualización del archivo")

            # Limitar a las primeras 50 filas para evitar sobrecarga
            df_preview = df_filtrado.head(50)

            styled_df = aplicar_estilo(df_preview)

            styled_html = styled_df.to_html()

            # Contenedor con scroll
            scrollable_html = f"""
            <div style="max-height: 600px; overflow-y: auto; border: 1px solid #ccc; padding: 10px">
                {styled_html}
            """

            st.markdown(scrollable_html, unsafe_allow_html=True)

            # Botón de descarga con todo el dataframe (sin limitar)
            excel_data = convertir_a_excel(df_filtrado)
            st.download_button(
                label="📥 Descargar Excel",
                data=excel_data,
                file_name= f"Formato_casos {hoy}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("No se ha cargado el DataFrame `df_filtrado`.")


